"""Report exporters: CSV, XLSX, MAT, PDF.

Closes one of the Kubios-parity gaps from docs/RUTHLESS_AUDIT_2026-04-21_CW.md
by providing the same four export formats Kubios HRV Premium ships
(PDF, Excel, MATLAB .mat, CSV). Each exporter takes an
AnalysisResponse and returns bytes; the route handler adds the
Content-Type + Content-Disposition headers and streams to the client.

The four formats serve different downstream readers:
  CSV  — pipeline-friendly, scripting, quick re-analysis in R/Python.
  XLSX — clinical / research colleague who lives in Excel.
  MAT  — MATLAB users (physiology labs, signal-processing courses).
  PDF  — paper appendix, lab notebook, IRB documentation.
"""

from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any

from app.schemas.analysis import AnalysisResponse


# ---------------------------------------------------------------------------
# Shared: flat (metric, value, unit, group) rows used by all exporters
# ---------------------------------------------------------------------------


def _summary_rows(analysis: AnalysisResponse) -> list[dict[str, Any]]:
    """Flatten AnalysisResponse into a list of rows for tabular formats.

    Each row has: metric (display name), value, unit, group. Groups
    follow the Kubios output-panel organisation:
      Time-Domain HRV / Frequency-Domain HRV / Poincaré / EDA /
      Sync QC / Metadata.
    """
    fs = analysis.feature_summary
    rows: list[dict[str, Any]] = []

    def add(metric: str, value: Any, unit: str, group: str) -> None:
        rows.append(
            {
                "metric": metric,
                "value": value,
                "unit": unit,
                "group": group,
            }
        )

    # Time-domain HRV ---------------------------------------------------------
    add("Mean HR", fs.mean_hr_bpm, "bpm", "Time-Domain HRV")
    add("RMSSD", fs.rmssd_ms, "ms", "Time-Domain HRV")
    add("SDNN", fs.sdnn_ms, "ms", "Time-Domain HRV")
    add("NN50", fs.nn50, "beats", "Time-Domain HRV")
    add("pNN50", fs.pnn50, "%", "Time-Domain HRV")

    # Poincaré nonlinear ------------------------------------------------------
    add("SD1", fs.sd1_ms, "ms", "Poincaré")
    add("SD2", fs.sd2_ms, "ms", "Poincaré")
    add("SD1/SD2", fs.sd1_sd2_ratio, "ratio", "Poincaré")
    add("Ellipse area", fs.ellipse_area_ms2, "ms²", "Poincaré")

    # Frequency-domain HRV ----------------------------------------------------
    add("VLF power", fs.vlf_ms2, "ms²", "Frequency-Domain HRV")
    add("LF power", fs.lf_ms2, "ms²", "Frequency-Domain HRV")
    add("HF power", fs.hf_ms2, "ms²", "Frequency-Domain HRV")
    add("Total power", fs.total_power_ms2, "ms²", "Frequency-Domain HRV")
    add("LF/HF ratio", fs.lf_hf_ratio, "ratio", "Frequency-Domain HRV")
    add("LF_nu", fs.lf_nu, "n.u.", "Frequency-Domain HRV")
    add("HF_nu", fs.hf_nu, "n.u.", "Frequency-Domain HRV")
    add("VLF%", fs.vlf_pct, "%", "Frequency-Domain HRV")
    add("LF%", fs.lf_pct, "%", "Frequency-Domain HRV")
    add("HF%", fs.hf_pct, "%", "Frequency-Domain HRV")

    # EDA ---------------------------------------------------------------------
    add("Tonic SCL (mean)", fs.eda_mean_us, "µS", "EDA")
    add("Phasic index", fs.eda_phasic_index, "µS", "EDA")

    # Stress composite --------------------------------------------------------
    add("Stress composite (0–1)", fs.stress_score, "0–1", "Stress")
    add("RR source", fs.rr_source, "label", "Stress")

    # Sync QC -----------------------------------------------------------------
    add("Sync QC score", analysis.sync_qc_score, "0–100", "Sync QC")
    add("Sync QC band", analysis.sync_qc_band, "label", "Sync QC")
    add("Sync QC gate", analysis.sync_qc_gate, "label", "Sync QC")
    add("Drift slope", analysis.drift_slope, "unitless", "Sync QC")
    add("Drift intercept", analysis.drift_intercept_ms, "ms", "Sync QC")
    add("Drift segments", analysis.drift_segments, "count", "Sync QC")

    # Metadata ----------------------------------------------------------------
    add("Synchronized samples", analysis.synchronized_samples, "count", "Metadata")
    add("Movement artifact ratio", analysis.movement_artifact_ratio, "0–1", "Metadata")

    return rows


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def export_to_csv(analysis: AnalysisResponse) -> bytes:
    """Flat (metric, value, unit, group) CSV."""
    rows = _summary_rows(analysis)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["group", "metric", "value", "unit"])
    writer.writeheader()
    for r in rows:
        writer.writerow(
            {
                "group": r["group"],
                "metric": r["metric"],
                "value": "" if r["value"] is None else r["value"],
                "unit": r["unit"],
            }
        )
    # Quality flags at the bottom, one per row
    buf.write("\n# Quality flags\n")
    for flag in analysis.quality_flags:
        buf.write(f'"{flag}"\n')
    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# XLSX (multi-sheet workbook)
# ---------------------------------------------------------------------------


def export_to_xlsx(analysis: AnalysisResponse) -> bytes:
    """Multi-sheet Excel workbook: one sheet per Kubios output panel."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    # Remove the default sheet; we add our own.
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="1C5A4E")
    header_font = Font(bold=True, color="FFFFFF")

    rows = _summary_rows(analysis)
    groups: dict[str, list[dict[str, Any]]] = {}
    for r in rows:
        groups.setdefault(r["group"], []).append(r)

    # Summary sheet listing every group --------------------------------------
    summary = wb.create_sheet("Summary")
    summary.append(["Group", "Metric", "Value", "Unit"])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")
    for r in rows:
        summary.append(
            [r["group"], r["metric"], r["value"], r["unit"]]
        )
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 28
    summary.column_dimensions["C"].width = 18
    summary.column_dimensions["D"].width = 10

    # Per-group sheets for easier pivot / cross-reference --------------------
    for group_name, group_rows in groups.items():
        safe = group_name.replace("/", "-").replace("é", "e")[:31]
        ws = wb.create_sheet(safe)
        ws.append(["Metric", "Value", "Unit"])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for r in group_rows:
            ws.append([r["metric"], r["value"], r["unit"]])
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 10

    # Quality flags sheet -----------------------------------------------------
    qf = wb.create_sheet("Quality flags")
    qf.append(["Flag"])
    qf["A1"].fill = header_fill
    qf["A1"].font = header_font
    for flag in analysis.quality_flags:
        qf.append([flag])
    qf.column_dimensions["A"].width = 120

    # Non-diagnostic notice sheet (appears on every Kubios PDF appendix)
    nd = wb.create_sheet("Non-diagnostic notice")
    nd["A1"] = "This report is research output. It is not a medical device."
    nd["A2"] = analysis.non_diagnostic_notice
    nd["A1"].font = Font(bold=True)
    nd.column_dimensions["A"].width = 120

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# MATLAB .mat
# ---------------------------------------------------------------------------


def export_to_mat(analysis: AnalysisResponse) -> bytes:
    """MATLAB struct-like .mat file with top-level `analysis` struct."""
    from scipy.io import savemat

    fs = analysis.feature_summary
    # Build a dict that scipy.io.savemat will write as a MATLAB struct.
    # None values become NaN; strings stay strings.
    def _n(v: Any) -> Any:
        return float("nan") if v is None else v

    data = {
        "analysis": {
            "time_domain": {
                "mean_hr_bpm": _n(fs.mean_hr_bpm),
                "rmssd_ms": _n(fs.rmssd_ms),
                "sdnn_ms": _n(fs.sdnn_ms),
                "nn50": _n(fs.nn50),
                "pnn50": _n(fs.pnn50),
            },
            "poincare": {
                "sd1_ms": _n(fs.sd1_ms),
                "sd2_ms": _n(fs.sd2_ms),
                "sd1_sd2_ratio": _n(fs.sd1_sd2_ratio),
                "ellipse_area_ms2": _n(fs.ellipse_area_ms2),
            },
            "frequency_domain": {
                "vlf_ms2": _n(fs.vlf_ms2),
                "lf_ms2": _n(fs.lf_ms2),
                "hf_ms2": _n(fs.hf_ms2),
                "total_power_ms2": _n(fs.total_power_ms2),
                "lf_hf_ratio": _n(fs.lf_hf_ratio),
                "lf_nu": _n(fs.lf_nu),
                "hf_nu": _n(fs.hf_nu),
                "vlf_pct": _n(fs.vlf_pct),
                "lf_pct": _n(fs.lf_pct),
                "hf_pct": _n(fs.hf_pct),
            },
            "eda": {
                "eda_mean_us": _n(fs.eda_mean_us),
                "eda_phasic_index": _n(fs.eda_phasic_index),
            },
            "stress": {
                "stress_score": _n(fs.stress_score),
                "rr_source": fs.rr_source,
            },
            "sync_qc": {
                "score": _n(analysis.sync_qc_score),
                "band": analysis.sync_qc_band,
                "gate": analysis.sync_qc_gate,
                "failure_reasons": list(analysis.sync_qc_failure_reasons),
                "drift_slope": _n(analysis.drift_slope),
                "drift_intercept_ms": _n(analysis.drift_intercept_ms),
                "drift_segments": _n(analysis.drift_segments),
            },
            "metadata": {
                "synchronized_samples": _n(analysis.synchronized_samples),
                "movement_artifact_ratio": _n(analysis.movement_artifact_ratio),
                "non_diagnostic_notice": analysis.non_diagnostic_notice,
                "quality_flags": list(analysis.quality_flags),
            },
        }
    }

    buf = io.BytesIO()
    # long_field_names=True lets descriptive names through; oned_as='column'
    # matches MATLAB's default orientation for 1-D arrays.
    savemat(buf, data, long_field_names=True, oned_as="column")
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def export_to_pdf(analysis: AnalysisResponse, *, session_id: str | None = None) -> bytes:
    """Formatted one-to-two-page PDF report, paper-appendix style."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=LETTER,
        topMargin=0.6 * inch,
        bottomMargin=0.6 * inch,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
    )
    styles = getSampleStyleSheet()
    h1 = styles["Heading1"]
    h2 = styles["Heading2"]
    body = styles["BodyText"]
    caption = ParagraphStyle("caption", parent=body, fontSize=8, textColor=colors.grey)

    story: list[Any] = []
    story.append(Paragraph("Polar-EmotiBit HRV report", h1))
    meta_line = datetime.utcnow().strftime("Generated %Y-%m-%d %H:%M UTC")
    if session_id:
        meta_line += f" · session_id = {session_id}"
    story.append(Paragraph(meta_line, caption))
    story.append(Spacer(1, 0.2 * inch))

    rows = _summary_rows(analysis)
    groups: dict[str, list[dict[str, Any]]] = {}
    for r in rows:
        groups.setdefault(r["group"], []).append(r)

    group_order = [
        "Time-Domain HRV",
        "Poincaré",
        "Frequency-Domain HRV",
        "EDA",
        "Stress",
        "Sync QC",
        "Metadata",
    ]
    for g in group_order:
        if g not in groups:
            continue
        story.append(Paragraph(g, h2))
        table_data: list[list[Any]] = [["Metric", "Value", "Unit"]]
        for r in groups[g]:
            val = r["value"]
            if isinstance(val, float):
                val = f"{val:.3f}"
            elif val is None:
                val = "—"
            table_data.append([r["metric"], str(val), r["unit"]])
        tbl = Table(
            table_data,
            colWidths=[2.6 * inch, 1.6 * inch, 0.9 * inch],
        )
        tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C5A4E")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F5F5F5")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
                    ("ALIGN", (1, 1), (2, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        story.append(tbl)
        story.append(Spacer(1, 0.14 * inch))

    # Quality flags
    story.append(Paragraph("Quality flags", h2))
    for flag in analysis.quality_flags:
        story.append(Paragraph(f"• {flag}", body))
    story.append(Spacer(1, 0.14 * inch))

    # Non-diagnostic footer
    story.append(PageBreak())
    story.append(Paragraph("Non-diagnostic notice", h2))
    story.append(Paragraph(analysis.non_diagnostic_notice, body))

    doc.build(story)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Dispatch
# ---------------------------------------------------------------------------


MIME_TYPES: dict[str, str] = {
    "csv": "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "mat": "application/x-matlab-data",
    "pdf": "application/pdf",
}

EXPORTERS: dict[str, Any] = {
    "csv": export_to_csv,
    "xlsx": export_to_xlsx,
    "mat": export_to_mat,
    "pdf": export_to_pdf,
}
