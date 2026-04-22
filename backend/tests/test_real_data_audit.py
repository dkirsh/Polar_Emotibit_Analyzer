"""Regression tests locking in the 2026-04-21 ruthless-audit P0 fixes.

Data source: Welltory PPG dataset (github.com/Welltory/welltory-ppg-dataset,
CC0-1.0). Two subjects converted to the Polar_Emotibit schema and
committed under data/samples/welltory/.

Guards:
- F12: HRV reads from raw Polar, RMSSD within 1 ms of ground-truth.
- F1:  Synthetic input (no rr_ms) honestly reports derived_from_bpm.
- F2/F6: Empty and below-minimum inputs return 422, not 200.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from fastapi.testclient import TestClient

from app.main import app
from app.services.ingestion.synthetic import generate_synthetic_session
from app.services.processing.features import lipponen_tarvainen_correction
from app.services.processing.pipeline import (
    InsufficientDataError,
    MIN_BEATS_FOR_HRV,
    run_analysis,
)


REPO_ROOT = Path(__file__).resolve().parents[2]
WELLTORY_DIR = REPO_ROOT / "data" / "samples" / "welltory"


def _load_welltory(name: str) -> pd.DataFrame:
    path = WELLTORY_DIR / f"welltory_{name}_polar.csv"
    assert path.exists(), f"missing fixture {path}"
    return pd.read_csv(path)


def _ground_truth_hrv(rr: np.ndarray, *, correct_ectopics: bool = True) -> tuple[float, float, float]:
    """Ground-truth HRV from RR intervals.

    Post-2026-04-21 the pipeline applies Lipponen-Tarvainen (2019)
    adaptive ectopic correction by default. The ground truth therefore
    applies the same correction to match; pass correct_ectopics=False
    for the raw textbook formulas used in the pre-L&T audit.
    """
    if correct_ectopics:
        rr = lipponen_tarvainen_correction(rr)[0]
    rmssd = float(np.sqrt(np.mean(np.diff(rr) ** 2)))
    sdnn = float(np.std(rr, ddof=1))
    mean_hr = float(60000.0 / np.mean(rr))
    return rmssd, sdnn, mean_hr


# -- F12: HRV reads from raw Polar RR (real-data evidence) ------------------


@pytest.mark.parametrize(
    "name, rmssd_tol, sdnn_tol, hr_tol",
    [
        ("s01", 1.0, 1.0, 1.0),   # 100 beats at 55 bpm
        ("s05", 1.0, 1.0, 1.0),   # 101 beats at 74 bpm — previously failed
    ],
)
def test_welltory_hrv_matches_ground_truth(name, rmssd_tol, sdnn_tol, hr_tol):
    """Pipeline HRV on real Polar data is within 1 ms / 1 bpm of direct calc."""
    polar = _load_welltory(name)
    rr = polar["rr_ms"].to_numpy(dtype=float)
    gt_rmssd, gt_sdnn, gt_hr = _ground_truth_hrv(rr)

    duration_s = int(polar["timestamp_ms"].iloc[-1] / 1000) + 1
    emo, _ = generate_synthetic_session(seconds=duration_s)

    result = run_analysis(emo, polar)
    fs = result.feature_summary

    assert abs(fs.rmssd_ms - gt_rmssd) < rmssd_tol, (
        f"{name}: RMSSD {fs.rmssd_ms:.3f} vs GT {gt_rmssd:.3f}"
    )
    assert abs(fs.sdnn_ms - gt_sdnn) < sdnn_tol, (
        f"{name}: SDNN {fs.sdnn_ms:.3f} vs GT {gt_sdnn:.3f}"
    )
    assert abs(fs.mean_hr_bpm - gt_hr) < hr_tol, (
        f"{name}: HR {fs.mean_hr_bpm:.3f} vs GT {gt_hr:.3f}"
    )
    assert fs.rr_source == "native_polar"


# -- F1: synthetic input honestly reports derived_from_bpm ------------------


def test_synthetic_input_reports_derived_rr_source():
    emo, polar = generate_synthetic_session(seconds=120)
    assert "rr_ms" not in polar.columns, "synthetic polar must not have rr_ms"
    result = run_analysis(emo, polar)
    assert result.feature_summary.rr_source == "derived_from_bpm"
    warn_flags = [f for f in result.quality_flags if "BPM-derived" in f]
    assert warn_flags, (
        "must emit a reduced-accuracy warning when rr_ms is absent"
    )


# -- F2 + F6: minimum-sample guard returns 422 via HTTP ---------------------


@pytest.fixture
def client():
    return TestClient(app)


@pytest.fixture
def meta():
    return {
        "session_id": "audit-regression",
        "subject_id": "test",
        "study_id": "polar-audit",
        "session_date": "2026-04-21",
    }


def test_empty_csv_returns_422(client, meta):
    r = client.post(
        "/api/v1/analyze",
        data=meta,
        files={
            "emotibit_file": ("e.csv", b"timestamp_ms,eda_us\n"),
            "polar_file": ("p.csv", b"timestamp_ms,hr_bpm\n"),
        },
    )
    assert r.status_code == 422
    body = r.json()
    assert body["detail"]["reason"] == "insufficient_data"
    assert body["detail"]["n_polar"] == 0


def test_below_minimum_beats_returns_422(client, meta):
    """49 beats should 422 (below MIN_BEATS_FOR_HRV=50)."""
    rr_values = [800] * 49
    polar_rows = "\n".join(
        f"{sum(rr_values[:i+1])},{60000/rr_values[i]:.2f},{rr_values[i]}"
        for i in range(len(rr_values))
    )
    polar_csv = f"timestamp_ms,hr_bpm,rr_ms\n{polar_rows}\n".encode()
    emo_rows = "\n".join(f"{i*1000},{2.5+0.01*i}" for i in range(40))
    emo_csv = f"timestamp_ms,eda_us\n{emo_rows}\n".encode()
    r = client.post(
        "/api/v1/analyze",
        data=meta,
        files={
            "emotibit_file": ("e.csv", emo_csv),
            "polar_file": ("p.csv", polar_csv),
        },
    )
    assert r.status_code == 422
    assert r.json()["detail"]["reason"] == "insufficient_data"


def test_at_minimum_beats_returns_200(client, meta):
    """50 beats (at threshold) should succeed."""
    rr_values = [800] * 50
    polar_rows = "\n".join(
        f"{sum(rr_values[:i+1])},{60000/rr_values[i]:.2f},{rr_values[i]}"
        for i in range(len(rr_values))
    )
    polar_csv = f"timestamp_ms,hr_bpm,rr_ms\n{polar_rows}\n".encode()
    emo_rows = "\n".join(f"{i*1000},{2.5+0.01*i}" for i in range(50))
    emo_csv = f"timestamp_ms,eda_us\n{emo_rows}\n".encode()
    r = client.post(
        "/api/v1/analyze",
        data=meta,
        files={
            "emotibit_file": ("e.csv", emo_csv),
            "polar_file": ("p.csv", polar_csv),
        },
    )
    assert r.status_code == 200, f"expected 200 at threshold, got {r.status_code}: {r.text[:300]}"


def test_below_minimum_emotibit_samples_returns_422(client, meta):
    """50 Polar beats but only 10 EmotiBit samples — should 422 on sync guard."""
    rr_values = [800] * 50
    polar_rows = "\n".join(
        f"{sum(rr_values[:i+1])},{60000/rr_values[i]:.2f},{rr_values[i]}"
        for i in range(len(rr_values))
    )
    polar_csv = f"timestamp_ms,hr_bpm,rr_ms\n{polar_rows}\n".encode()
    # Only 10 EmotiBit samples, below MIN_SAMPLES_FOR_SYNC = 30
    emo_rows = "\n".join(f"{i*1000},{2.5+0.01*i}" for i in range(10))
    emo_csv = f"timestamp_ms,eda_us\n{emo_rows}\n".encode()
    r = client.post(
        "/api/v1/analyze",
        data=meta,
        files={
            "emotibit_file": ("e.csv", emo_csv),
            "polar_file": ("p.csv", polar_csv),
        },
    )
    assert r.status_code == 422
    assert r.json()["detail"]["reason"] == "insufficient_data"
    assert r.json()["detail"]["n_emotibit"] == 10


# -- Kubios-parity: four export formats on real Welltory data ---------------


def _welltory_analysis(name: str = "s05"):
    """Run the pipeline on a Welltory sample and return the AnalysisResponse."""
    polar = _load_welltory(name)
    duration_s = int(polar["timestamp_ms"].iloc[-1] / 1000) + 1
    emo, _ = generate_synthetic_session(seconds=duration_s)
    return run_analysis(emo, polar)


def test_export_csv_is_parseable():
    from app.services.reporting.exporters import export_to_csv
    analysis = _welltory_analysis()
    out = export_to_csv(analysis)
    text = out.decode("utf-8")
    header = text.splitlines()[0]
    assert header == "group,metric,value,unit"
    # Quality flags section must be present
    assert "# Quality flags" in text


def test_export_xlsx_has_expected_sheets():
    import io as _io
    from openpyxl import load_workbook
    from app.services.reporting.exporters import export_to_xlsx
    analysis = _welltory_analysis()
    out = export_to_xlsx(analysis)
    wb = load_workbook(_io.BytesIO(out))
    assert "Summary" in wb.sheetnames
    assert "Quality flags" in wb.sheetnames
    assert "Non-diagnostic notice" in wb.sheetnames
    # At least one group sheet
    group_sheets = [s for s in wb.sheetnames if s in {
        "Time-Domain HRV", "Poincare", "Frequency-Domain HRV", "EDA",
    }]
    assert len(group_sheets) >= 3


def test_export_mat_has_analysis_struct():
    import io as _io
    from scipy.io import loadmat
    from app.services.reporting.exporters import export_to_mat
    analysis = _welltory_analysis()
    out = export_to_mat(analysis)
    mat = loadmat(_io.BytesIO(out), squeeze_me=True, struct_as_record=False)
    assert "analysis" in mat
    ana = mat["analysis"]
    # The time_domain sub-struct must carry the core time-domain metrics
    td = ana.time_domain
    assert hasattr(td, "rmssd_ms")
    assert hasattr(td, "sdnn_ms")
    assert hasattr(td, "mean_hr_bpm")


def test_export_pdf_has_pdf_header():
    from app.services.reporting.exporters import export_to_pdf
    analysis = _welltory_analysis()
    out = export_to_pdf(analysis, session_id="welltory-s05")
    # PDF files begin with %PDF-1.x
    assert out.startswith(b"%PDF-"), f"not a PDF; first 8 bytes = {out[:8]!r}"
    assert len(out) > 1000  # non-trivial content


def test_export_endpoint_roundtrip_404_on_missing():
    """GET /api/v1/sessions/unknown/export?format=csv -> 404."""
    c = TestClient(app)
    r = c.get("/api/v1/sessions/does-not-exist/export", params={"format": "csv"})
    assert r.status_code == 404


def test_export_endpoint_rejects_unknown_format():
    """Unknown format returns 422."""
    # Need a stored session; import from the route module
    from app.api.v1.routes.analysis import _SESSION_STORE
    _SESSION_STORE["fmt-test"] = {
        "session_id": "fmt-test",
        "subject_id": "x", "study_id": "y", "session_date": "2026-04-21",
        "analyzed_at": "2026-04-21T00:00:00",
        "result": _welltory_analysis().model_dump(),
    }
    c = TestClient(app)
    r = c.get("/api/v1/sessions/fmt-test/export", params={"format": "exe"})
    assert r.status_code == 422
    _SESSION_STORE.pop("fmt-test", None)
